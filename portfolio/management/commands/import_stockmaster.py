# portfolio/management/commands/import_stockmaster.py
import requests
from pathlib import Path
from django.core.management.base import BaseCommand
from portfolio.models import StockMaster
from openpyxl import load_workbook
from django.conf import settings

# 保存先を settings から取得
STOCKMASTER_XLSX_PATH = getattr(settings, "STOCKMASTER_XLSX_PATH", Path(settings.BASE_DIR) / "data" / "StockMaster_latest.xlsx")
JPX_XLSX_URL = "https://www.jpx.co.jp/markets/statistics-equities/misc/tse-listed-issues.xlsx"

class Command(BaseCommand):
    help = "JPX公式.xlsxから銘柄マスタ（StockMaster）を更新・追加"

    def handle(self, *args, **options):
        # 保存先ディレクトリ作成
        STOCKMASTER_XLSX_PATH.parent.mkdir(parents=True, exist_ok=True)

        # 1. Excel自動ダウンロード
        self.stdout.write("JPX公式Excelを自動ダウンロード中...")
        try:
            res = requests.get(JPX_XLSX_URL, timeout=15)
            res.raise_for_status()
            with open(STOCKMASTER_XLSX_PATH, "wb") as f:
                f.write(res.content)
            self.stdout.write(self.style.SUCCESS(f"Excelを保存: {STOCKMASTER_XLSX_PATH}"))
        except requests.RequestException as e:
            self.stdout.write(self.style.WARNING(f"Excelダウンロード失敗: {e}"))
            if not STOCKMASTER_XLSX_PATH.exists():
                self.stdout.write(self.style.ERROR("既存ファイルもないため終了"))
                return
            self.stdout.write("既存Excelを使用します…")

        # 2. Excel 読み込み
        try:
            wb = load_workbook(STOCKMASTER_XLSX_PATH, data_only=True)
            ws = wb.active
        except Exception as e:
            self.stdout.write(self.style.ERROR(f"Excel読み込み失敗: {e}"))
            return

        # 3. 先頭行をヘッダーとして読み込み
        headers = [cell.value for cell in ws[1]]
        code_idx = headers.index("証券コード") if "証券コード" in headers else headers.index("Code")
        name_idx = headers.index("銘柄名") if "銘柄名" in headers else headers.index("Name")
        sector_idx = headers.index("33業種") if "33業種" in headers else headers.index("Sector")

        # 4. データ更新
        created = updated = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            code = str(row[code_idx]).zfill(4)
            name = str(row[name_idx]).strip()
            sector = str(row[sector_idx]).strip()
            if not (code and name and sector):
                continue
            obj, created_flag = StockMaster.objects.update_or_create(
                code=code,
                defaults={"name": name, "sector": sector},
            )
            if created_flag:
                created += 1
            else:
                updated += 1

        self.stdout.write(self.style.SUCCESS(f"作成 {created} 件、更新 {updated} 件"))
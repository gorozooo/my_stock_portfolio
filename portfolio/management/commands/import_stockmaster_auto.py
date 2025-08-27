# portfolio/management/commands/import_stockmaster_auto.py
import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from django.core.management.base import BaseCommand
from portfolio.models import StockMaster
from io import BytesIO

JPX_URL = "https://www.jpx.co.jp/markets/statistics-equities/misc/"

class Command(BaseCommand):
    help = "JPX 上場銘柄一覧を自動で取得・更新します"

    def handle(self, *args, **options):
        self.stdout.write("JPX 上場銘柄一覧ページを取得中...")
        try:
            res = requests.get(JPX_URL)
            res.raise_for_status()
        except Exception as e:
            self.stderr.write(f"ページ取得失敗: {e}")
            return

        soup = BeautifulSoup(res.text, "html.parser")

        # Excelリンクを探す
        link = None
        for a in soup.select("a"):
            href = a.get("href", "")
            if "tse-listed-issues.xlsx" in href:
                link = href
                break

        if not link:
            self.stderr.write("Excelリンクが見つかりません")
            return

        # 完全URLに変換
        if not link.startswith("http"):
            link = "https://www.jpx.co.jp" + link

        self.stdout.write(f"Excelをダウンロード中: {link}")
        try:
            excel_res = requests.get(link)
            excel_res.raise_for_status()
        except Exception as e:
            self.stderr.write(f"Excelダウンロードに失敗しました: {e}")
            return

        try:
            wb = load_workbook(filename=BytesIO(excel_res.content), data_only=True)
        except Exception as e:
            self.stderr.write(f"Excel読み込みに失敗しました: {e}")
            return

        ws = wb.active
        count = 0

        for row in ws.iter_rows(min_row=2, values_only=True):
            code = str(row[0]).zfill(4) if row[0] else None
            name = row[1] if len(row) > 1 else None
            sector = row[2] if len(row) > 2 else ""

            if not code or not name:
                continue

            StockMaster.objects.update_or_create(
                code=code,
                defaults={"name": name, "sector": sector}
            )
            count += 1

        self.stdout.write(self.style.SUCCESS(f"{count}件の銘柄を更新/追加しました"))
